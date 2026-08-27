"""
PLATAFORMA WEB - OPTIMIZACIÓN DE RUTAS TURÍSTICAS
Provincia de Coclé, Panamá
28 Atractivos | 7 Itinerarios | Algoritmo de Dijkstra
TODO EN UN SOLO ARCHIVO - SIN IMPORTACIONES EXTERNAS
"""

import os
import heapq
from itertools import permutations
from io import BytesIO
import streamlit as st
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# 1. ATRACTIVOS TURÍSTICOS (28 NODOS)
# ═══════════════════════════════════════════════════════════════════════════

ATRACTIVOS = {
    # === PLAYAS (Costa Pacífica) ===
    1: {"nombre": "Playa Santa Clara", "cod": "PSC", "tipo": "Playa", "puntaje": 27, "distrito": "Antón"},
    2: {"nombre": "Playa Farallón", "cod": "PFA", "tipo": "Playa", "puntaje": 26, "distrito": "Antón"},
    3: {"nombre": "Playa El Salado", "cod": "PES", "tipo": "Playa", "puntaje": 21, "distrito": "Aguadulce"},
    4: {"nombre": "Playa Blanca", "cod": "PBL", "tipo": "Playa", "puntaje": 26, "distrito": "Antón"},
    5: {"nombre": "Playa Juan Hombrón", "cod": "PJH", "tipo": "Playa", "puntaje": 20, "distrito": "Antón"},
    26: {"nombre": "Playa La Hueca", "cod": "PLH", "tipo": "Playa", "puntaje": 22, "distrito": "Antón"},
    
    # === CULTURA E HISTORIA ===
    6: {"nombre": "Mercado Artesanía Valle Antón", "cod": "MAV", "tipo": "Cultural", "puntaje": 26, "distrito": "Antón"},
    8: {"nombre": "Museo Hermanos Arias Madrid", "cod": "MHA", "tipo": "Cultural/Hist.", "puntaje": 25, "distrito": "Penonomé"},
    11: {"nombre": "Museo Regional Stella Sierra", "cod": "MSS", "tipo": "Cultural/Hist.", "puntaje": 22, "distrito": "Aguadulce"},
    25: {"nombre": "Museo de la Sal", "cod": "MSA", "tipo": "Cultural", "puntaje": 21, "distrito": "Aguadulce"},
    12: {"nombre": "Iglesia San Juan Bautista", "cod": "ISJ", "tipo": "Histórico", "puntaje": 24, "distrito": "Penonomé"},
    20: {"nombre": "Parroquia Ntra. Sra. Candelaria", "cod": "PNC", "tipo": "Histórico", "puntaje": 21, "distrito": "La Pintada"},
    
    # === NATURALEZA Y PARQUES ===
    7: {"nombre": "Serpentario Maravillas Tropicales", "cod": "SMT", "tipo": "Naturaleza", "puntaje": 24, "distrito": "Antón"},
    9: {"nombre": "P.N. Omar Torrijos", "cod": "PNT", "tipo": "Parque Nacional", "puntaje": 22, "distrito": "Penonomé"},
    10: {"nombre": "Sitio Arqueológico El Caño", "cod": "SAC", "tipo": "Arqueológico", "puntaje": 26, "distrito": "Natá"},
    13: {"nombre": "El Chorro Las Yayas", "cod": "CLY", "tipo": "Cascada", "puntaje": 25, "distrito": "La Pintada"},
    14: {"nombre": "Balneario Las Mendozas", "cod": "BLM", "tipo": "Balneario", "puntaje": 21, "distrito": "Penonomé"},
    21: {"nombre": "Cerro Gaital", "cod": "CGA", "tipo": "Montaña", "puntaje": 27, "distrito": "Antón"},
    22: {"nombre": "Manglares de Aguadulce", "cod": "MAG", "tipo": "Naturaleza", "puntaje": 23, "distrito": "Aguadulce"},
    23: {"nombre": "Cerro El Valle", "cod": "CEV", "tipo": "Montaña", "puntaje": 24, "distrito": "Antón"},
    24: {"nombre": "Balneario El Copé", "cod": "BEC", "tipo": "Balneario", "puntaje": 22, "distrito": "La Pintada"},
    27: {"nombre": "Cerro La Cruz", "cod": "CLC", "tipo": "Montaña", "puntaje": 23, "distrito": "Penonomé"},
    28: {"nombre": "Mirador de Natá", "cod": "MIN", "tipo": "Mirador", "puntaje": 22, "distrito": "Natá"},
    
    # === CIUDADES / HUBS ===
    15: {"nombre": "Penonomé", "cod": "PEN", "tipo": "Hub/Ciudad", "puntaje": 28, "distrito": "Penonomé"},
    16: {"nombre": "Aguadulce", "cod": "AGU", "tipo": "Hub/Ciudad", "puntaje": 25, "distrito": "Aguadulce"},
    17: {"nombre": "Antón", "cod": "ANT", "tipo": "Hub/Ciudad", "puntaje": 23, "distrito": "Antón"},
    18: {"nombre": "La Pintada", "cod": "LAP", "tipo": "Hub/Ciudad", "puntaje": 22, "distrito": "La Pintada"},
    19: {"nombre": "Natá", "cod": "NAT", "tipo": "Hub/Ciudad", "puntaje": 23, "distrito": "Natá"},
}

# ═══════════════════════════════════════════════════════════════════════════
# 2. ARISTAS - CONEXIONES VIALES (44 ARISTAS)
# ═══════════════════════════════════════════════════════════════════════════

ARISTAS = [
    # === CONEXIONES PRINCIPALES (Panamericana y rutas principales) ===
    (15, 17, 35.0, 35),   # Penonomé ↔ Antón
    (15, 16, 52.2, 48),   # Penonomé ↔ Aguadulce
    (15, 18, 26.0, 30),   # Penonomé ↔ La Pintada
    (15, 19, 67.0, 67),   # Penonomé ↔ Natá
    (15, 8, 2.5, 5),      # Penonomé ↔ Museo Hnos. Arias
    (15, 12, 2.0, 4),     # Penonomé ↔ Iglesia San Juan
    (15, 14, 3.0, 6),     # Penonomé ↔ Balneario Mendozas
    (15, 21, 42.6, 40),   # Penonomé ↔ Cerro Gaital
    (15, 10, 78.0, 55),   # Penonomé ↔ El Caño
    (19, 10, 8.0, 12),    # Natá ↔ El Caño
    (19, 16, 28.0, 32),   # Natá ↔ Aguadulce
    (19, 17, 32.0, 35),   # Natá ↔ Antón
    (16, 10, 26.8, 27),   # Aguadulce ↔ El Caño
    (16, 11, 5.0, 8),     # Aguadulce ↔ Museo Stella Sierra
    (16, 3, 8.5, 12),     # Aguadulce ↔ Playa El Salado
    (18, 9, 35.0, 50),    # La Pintada ↔ Omar Torrijos
    (18, 13, 14.0, 20),   # La Pintada ↔ El Chorro Las Yayas
    (18, 20, 0.5, 2),     # La Pintada ↔ Parroquia Candelaria
    (17, 1, 18.0, 20),    # Antón ↔ Playa Santa Clara
    (17, 2, 20.5, 22),    # Antón ↔ Playa Farallón
    (17, 4, 22.0, 24),    # Antón ↔ Playa Blanca
    (17, 5, 14.5, 16),    # Antón ↔ Playa Juan Hombrón
    (17, 6, 22.0, 25),    # Antón ↔ Mercado Artesanía
    (17, 7, 22.5, 26),    # Antón ↔ Serpentario
    (17, 21, 10.0, 15),   # Antón ↔ Cerro Gaital
    (1, 2, 6.8, 11),      # P. Santa Clara ↔ P. Farallón
    (2, 4, 2.5, 5),       # P. Farallón ↔ P. Blanca
    (4, 5, 9.5, 12),      # P. Blanca ↔ P. Juan Hombrón
    (1, 5, 17.0, 20),     # P. Santa Clara ↔ P. Juan Hombrón
    (6, 7, 0.5, 2),       # Mercado ↔ Serpentario
    (6, 21, 38.0, 45),    # Mercado ↔ Cerro Gaital
    (9, 13, 12.0, 18),    # Omar Torrijos ↔ El Chorro
    (9, 10, 39.7, 65),    # Omar Torrijos ↔ El Caño
    
    # === NUEVAS ARISTAS (Nodos 22-28) ===
    (16, 22, 3.5, 5),     # Aguadulce ↔ Manglares
    (16, 25, 2.0, 4),     # Aguadulce ↔ Museo de la Sal
    (22, 25, 1.5, 3),     # Manglares ↔ Museo de la Sal
    (17, 23, 8.0, 12),    # Antón ↔ Cerro El Valle
    (17, 26, 15.0, 18),   # Antón ↔ Playa La Hueca
    (23, 26, 12.0, 15),   # Cerro El Valle ↔ Playa La Hueca
    (18, 24, 12.0, 18),   # La Pintada ↔ Balneario El Copé
    (24, 9, 18.0, 25),    # El Copé ↔ Omar Torrijos
    (15, 27, 3.0, 6),     # Penonomé ↔ Cerro La Cruz
    (19, 28, 2.5, 5),     # Natá ↔ Mirador de Natá
    (28, 10, 6.0, 10),    # Mirador ↔ El Caño
]

# ═══════════════════════════════════════════════════════════════════════════
# 3. CONFIGURACIÓN DE DÍAS (7 RUTAS CIRCULARES)
# ═══════════════════════════════════════════════════════════════════════════

DIAS_CONFIG = {
    1: {"destinos": [1, 2, 4, 5], "hub": 17, "zona": "Ruta Costera Norte – Playas de Antón", "color": "#185FA5"},
    2: {"destinos": [6, 7, 23, 21], "hub": 17, "zona": "Valle de Antón y Montañas", "color": "#854F0B"},
    3: {"destinos": [8, 12, 27, 14], "hub": 15, "zona": "Penonomé Cultural y Natural", "color": "#0F6E56"},
    4: {"destinos": [13, 24, 9, 20], "hub": 18, "zona": "Circuito Montañoso – Cascadas y Parque", "color": "#534AB7"},
    5: {"destinos": [3, 25, 22, 11, 16], "hub": 16, "zona": "Aguadulce – Costa Sur y Cultura", "color": "#993C1D"},
    6: {"destinos": [28, 10, 19], "hub": 19, "zona": "Zona Arqueológica – El Caño y Natá", "color": "#0F6E56"},
    7: {"destinos": [26, 5, 17, 19, 28], "hub": 15, "zona": "Circuito Integrador – Costa, Hubs y Miradores", "color": "#5B4A00"},
}

# ═══════════════════════════════════════════════════════════════════════════
# 4. COLORES Y POSICIONES
# ═══════════════════════════════════════════════════════════════════════════

COLORES_TIPO = {
    "Hub/Ciudad": "#1D9E75",
    "Playa": "#378ADD",
    "Cultural": "#BA7517",
    "Cultural/Hist.": "#BA7517",
    "Histórico": "#BA7517",
    "Naturaleza": "#7F77DD",
    "Parque Nacional": "#7F77DD",
    "Cascada": "#7F77DD",
    "Balneario": "#7F77DD",
    "Arqueológico": "#D85A30",
    "Montaña": "#7F77DD",
    "Mirador": "#FF6B35",
}

POSICIONES = {
    15: (0.50, 0.50), 17: (0.74, 0.42), 16: (0.22, 0.20), 18: (0.36, 0.74),
    19: (0.40, 0.26), 1: (0.94, 0.62), 2: (0.94, 0.50), 3: (0.10, 0.10),
    4: (0.92, 0.38), 5: (0.84, 0.26), 6: (0.82, 0.68), 7: (0.94, 0.74),
    8: (0.44, 0.44), 9: (0.18, 0.82), 10: (0.30, 0.30), 11: (0.10, 0.24),
    12: (0.58, 0.44), 13: (0.20, 0.90), 14: (0.50, 0.36), 20: (0.28, 0.84),
    21: (0.80, 0.82), 22: (0.15, 0.15), 23: (0.88, 0.78), 24: (0.28, 0.78),
    25: (0.18, 0.18), 26: (0.90, 0.20), 27: (0.56, 0.56), 28: (0.34, 0.28),
}

# ═══════════════════════════════════════════════════════════════════════════
# 5. ALGORITMO DE DIJKSTRA
# ═══════════════════════════════════════════════════════════════════════════

def construir_grafo():
    G = nx.Graph()
    for nid, data in ATRACTIVOS.items():
        G.add_node(nid, **data)
    for u, v, dist, tiempo in ARISTAS:
        costo = round(dist * 0.15, 2)
        G.add_edge(u, v, distancia=dist, tiempo=tiempo, costo=costo)
    return G


def dijkstra(G, origen, criterio):
    INF = float('inf')
    dist = {n: INF for n in G.nodes()}
    prev = {n: None for n in G.nodes()}
    dist[origen] = 0
    heap = [(0, origen)]
    visitados = set()

    while heap:
        d_u, u = heapq.heappop(heap)
        if u in visitados:
            continue
        visitados.add(u)
        for v in G.neighbors(u):
            peso = G[u][v][criterio]
            alt = dist[u] + peso
            if alt < dist[v]:
                dist[v] = alt
                prev[v] = u
                heapq.heappush(heap, (alt, v))
    return dist, prev


def reconstruir_camino(prev, origen, destino):
    camino = []
    actual = destino
    while actual is not None:
        camino.append(actual)
        actual = prev[actual]
    camino.reverse()
    return camino if camino and camino[0] == origen else []


def calcular_todas_matrices(G):
    nodos = sorted(G.nodes())
    matrices, caminos = {}, {}
    for criterio in ["distancia", "tiempo", "costo"]:
        mat, cam = {}, {}
        for origen in nodos:
            dist_min, prev = dijkstra(G, origen, criterio)
            mat[origen] = {d: round(dist_min[d], 2) if dist_min[d] != float('inf') else None
                          for d in nodos}
            cam[origen] = {d: reconstruir_camino(prev, origen, d) for d in nodos}
        matrices[criterio] = mat
        caminos[criterio] = cam
    return matrices, caminos


def ruta_optima_dia(destinos, matrices, hub=None, criterio="tiempo"):
    if hub is None:
        hub = 15
    mat = matrices[criterio]
    mejor_costo = float('inf')
    mejor_orden = None

    for perm in permutations(destinos):
        secuencia = [hub] + list(perm) + [hub]
        total, valida = 0, True
        for i in range(len(secuencia) - 1):
            c = mat[secuencia[i]][secuencia[i + 1]]
            if c is None or c == float('inf'):
                valida = False
                break
            total += c
        if valida and total < mejor_costo:
            mejor_costo = total
            mejor_orden = list(perm)

    return mejor_orden, mejor_costo


# ═══════════════════════════════════════════════════════════════════════════
# 6. EXPORTACIÓN A EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def generar_excel_bytes(matrices, G):
    wb = Workbook()
    wb.remove(wb.active)

    AZUL, VERDE, NARAN, GRIS, BLANC = "1F4E79", "1D9E75", "C55A11", "F2F2F2", "FFFFFF"

    def celda(ws, fila, col, valor, bold=False, bg=None, color="000000", alinear="center", size=10):
        c = ws.cell(row=fila, column=col, value=valor)
        c.font = Font(name="Arial", bold=bold, color=color, size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=alinear, vertical="center", wrap_text=True)
        lado = Side(style="thin", color="AAAAAA")
        c.border = Border(left=lado, right=lado, top=lado, bottom=lado)
        return c

    config = [
        ("Matriz_Tiempo", "tiempo", "MATRIZ DIJKSTRA – TIEMPO MÍNIMO (minutos)", AZUL),
        ("Matriz_Distancia", "distancia", "MATRIZ DIJKSTRA – DISTANCIA MÍNIMA (km)", VERDE),
        ("Matriz_Costo", "costo", "MATRIZ DIJKSTRA – COSTO MÍNIMO (USD)", NARAN),
    ]

    nodos = sorted(G.nodes())
    for sheet, criterio, titulo, color in config:
        ws = wb.create_sheet(sheet)
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(nodos)+1)
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = Font(name="Arial", bold=True, color=BLANC, size=12)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        celda(ws, 2, 1, "O \\ D", bold=True, bg=GRIS, size=8)
        for j, n in enumerate(nodos, 2):
            celda(ws, 2, j, f"{n}\n{ATRACTIVOS[n]['cod']}", bold=True, bg=GRIS, size=7)
            ws.column_dimensions[get_column_letter(j)].width = 9
        ws.column_dimensions["A"].width = 12
        ws.row_dimensions[2].height = 28

        mat = matrices[criterio]
        for i, origen in enumerate(nodos, 3):
            celda(ws, i, 1, f"{origen} {ATRACTIVOS[origen]['cod']}", bold=True, bg=GRIS, size=8)
            ws.row_dimensions[i].height = 18
            for j, destino in enumerate(nodos, 2):
                if origen == destino:
                    celda(ws, i, j, 0, bg="D9D9D9", size=8)
                else:
                    val = mat[origen][destino]
                    if val is None:
                        celda(ws, i, j, "∞", bg="FFE2CC", size=8)
                    else:
                        fmt = f"${val:.2f}" if criterio == "costo" else round(val, 1)
                        bg = GRIS if i % 2 == 0 else BLANC
                        celda(ws, i, j, fmt, bg=bg, size=8)
        ws.freeze_panes = "B3"

    # Inventario
    ws_inv = wb.create_sheet("Inventario_Atractivos")
    ws_inv.sheet_view.showGridLines = False
    ws_inv.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws_inv.cell(row=1, column=1, value="INVENTARIO DE ATRACTIVOS TURÍSTICOS – COCLÉ, PANAMÁ")
    c.font = Font(name="Arial", bold=True, color=BLANC, size=12)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_inv.row_dimensions[1].height = 26

    hdrs = ["ID", "Código", "Nombre", "Tipo", "Distrito", "Puntaje", "Grado", "Tipo Hub"]
    for j, h in enumerate(hdrs, 1):
        celda(ws_inv, 2, j, h, bold=True, bg="BDD7EE", size=10)

    for i, (nid, data) in enumerate(ATRACTIVOS.items(), 3):
        bg = GRIS if i % 2 == 0 else BLANC
        grado = G.degree(nid)
        es_hub = "✔" if data["tipo"] == "Hub/Ciudad" else ""
        for j, val in enumerate([nid, data["cod"], data["nombre"], data["tipo"],
                                  data["distrito"], data["puntaje"], grado, es_hub], 1):
            al = "left" if j == 3 else "center"
            celda(ws_inv, i, j, val, bg=bg, alinear=al, size=10)

    anchos = [5, 7, 38, 20, 14, 9, 8, 9]
    for j, w in enumerate(anchos, 1):
        ws_inv.column_dimensions[get_column_letter(j)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════════
# 7. CONFIGURACIÓN DE STREAMLIT
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Rutas Turísticas Coclé — 28 Atractivos",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded"
)

CARPETA_IMAGENES = os.path.join(os.path.dirname(__file__), "imagenes")

# ═══════════════════════════════════════════════════════════════════════════
# 8. CARGAR DATOS (CACHE)
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_resource
def cargar_datos():
    G = construir_grafo()
    matrices, caminos = calcular_todas_matrices(G)
    return G, matrices, caminos


G, matrices, caminos = cargar_datos()

UNIDAD = {"distancia": "km", "tiempo": "min", "costo": "USD"}
UNIDAD_SYM = {"distancia": "km", "tiempo": "min", "costo": "$"}
NOMBRE_CRIT = {"distancia": "Distancia mínima (km)", "tiempo": "Tiempo mínimo (min)", "costo": "Costo mínimo (USD)"}


# ═══════════════════════════════════════════════════════════════════════════
# 9. FUNCIONES DE DIBUJO
# ═══════════════════════════════════════════════════════════════════════════

def dibujar_grafo_completo(criterio):
    fig, ax = plt.subplots(figsize=(16, 12))
    ax.set_facecolor("#F0F4F8")
    fig.patch.set_facecolor("#E8EDF2")

    for u, v, data in G.edges(data=True):
        x0, y0 = POSICIONES[u]; x1, y1 = POSICIONES[v]
        ax.plot([x0, x1], [y0, y1], color="#BBBBBB", linewidth=1.0, zorder=1, alpha=0.6)
        mx, my = (x0 + x1) / 2, (y0 + y1) / 2
        val = data[criterio]
        txt = f"${val:.2f}" if criterio == "costo" else f"{val}{UNIDAD_SYM[criterio]}"
        ax.text(mx, my, txt, fontsize=6, ha="center", va="center", color="#555555", zorder=2,
                bbox=dict(boxstyle="round,pad=0.15", facecolor="white", edgecolor="#DDDDDD", alpha=0.85, linewidth=0.5))

    RADIO_HUB, RADIO_NORM = 0.035, 0.025
    for nid in G.nodes():
        x, y = POSICIONES[nid]
        tipo = G.nodes[nid]["tipo"]
        color = COLORES_TIPO.get(tipo, "#999999")
        radio = RADIO_HUB if tipo == "Hub/Ciudad" else RADIO_NORM
        ax.add_patch(plt.Circle((x + 0.004, y - 0.004), radio, color="#00000022", zorder=3))
        ax.add_patch(plt.Circle((x, y), radio, color=color, zorder=4, linewidth=2.0, ec="white"))
        ax.text(x, y + 0.006, G.nodes[nid]["cod"], ha="center", va="center", fontsize=8.5,
                fontweight="bold", color="white", zorder=5)
        nombre = G.nodes[nid]["nombre"]
        if x > 0.78:
            lx, ly, ha = x + 0.008, y + radio + 0.024, "left"
        elif x < 0.22:
            lx, ly, ha = x - 0.008, y + radio + 0.024, "right"
        elif y > 0.68:
            lx, ly, ha = x, y + radio + 0.024, "center"
        else:
            lx, ly, ha = x, y - radio - 0.024, "center"
        ax.text(lx, ly, nombre, ha=ha, va="center", fontsize=7, fontweight="600", color="#1A1A1A", zorder=6,
                bbox=dict(boxstyle="round,pad=0.25", facecolor="white", edgecolor=color, alpha=0.93, linewidth=0.8))

    leyenda = [
        mpatches.Patch(color="#1D9E75", label="Hub / Ciudad"),
        mpatches.Patch(color="#378ADD", label="Playa"),
        mpatches.Patch(color="#BA7517", label="Cultural / Histórico"),
        mpatches.Patch(color="#7F77DD", label="Naturaleza / Parque / Montaña"),
        mpatches.Patch(color="#D85A30", label="Arqueológico"),
        mpatches.Patch(color="#FF6B35", label="Mirador"),
    ]
    ax.legend(handles=leyenda, loc="lower left", fontsize=9, framealpha=0.95, edgecolor="#AAAAAA", fancybox=True)
    ax.set_title(f"Grafo Turístico G=(V,E) — Coclé, Panamá\n28 Nodos · {G.number_of_edges()} Aristas | Criterio: {NOMBRE_CRIT[criterio]}",
                 fontsize=13, fontweight="bold", color="#1F4E79", pad=14)
    ax.set_xlim(-0.08, 1.12); ax.set_ylim(-0.06, 1.06)
    ax.axis("off")
    plt.tight_layout(pad=1.5)
    return fig


def dibujar_grafo_dia(dia, criterio, secuencia, costo_total):
    cfg = DIAS_CONFIG[dia]
    destinos = cfg["destinos"]
    hub = cfg["hub"]
    color_dia = cfg["color"]

    nodos_ruta = set(secuencia)
    aristas_ruta, segmentos_dir = [], []
    for i in range(len(secuencia) - 1):
        seg = caminos[criterio][secuencia[i]][secuencia[i + 1]]
        for j in range(len(seg) - 1):
            aristas_ruta.append(tuple(sorted([seg[j], seg[j + 1]])))
            segmentos_dir.append((seg[j], seg[j + 1]))
        nodos_ruta.update(seg)
    aristas_ruta = list(set(aristas_ruta))

    nodos_destino = set(destinos)
    nodos_transito = nodos_ruta - nodos_destino - {hub}
    pos_sub = {n: POSICIONES[n] for n in nodos_ruta}

    fig, ax = plt.subplots(figsize=(14, 11))
    ax.set_facecolor("#F0F4F8")
    fig.patch.set_facecolor("#E8EDF2")

    for u, v in G.edges():
        if u in nodos_ruta and v in nodos_ruta and tuple(sorted([u, v])) not in aristas_ruta:
            x0, y0 = pos_sub[u]; x1, y1 = pos_sub[v]
            ax.plot([x0, x1], [y0, y1], color="#CCCCCC", linewidth=0.6, alpha=0.3, zorder=1, linestyle="dotted")

    ya_etiquetado = set()
    for u, v in segmentos_dir:
        x0, y0 = pos_sub[u]; x1, y1 = pos_sub[v]
        ax.annotate("", xy=(x1, y1), xytext=(x0, y0),
                    arrowprops=dict(arrowstyle="-|>", color=color_dia, lw=2.8,
                                     connectionstyle="arc3,rad=0.05", mutation_scale=18), zorder=3)
        par = tuple(sorted([u, v]))
        if par not in ya_etiquetado and G.has_edge(u, v):
            val = G[u][v][criterio]
            txt = f"${val:.2f}" if criterio == "costo" else f"{val}{UNIDAD_SYM[criterio]}"
            mx, my = (x0 + x1) / 2, (y0 + y1) / 2
            ax.text(mx, my, txt, ha="center", va="center", fontsize=8, fontweight="bold", color=color_dia, zorder=5,
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor=color_dia, alpha=0.95, linewidth=1.0))
            ya_etiquetado.add(par)

    RADIO_BASE, RADIO_DEST, RADIO_TRANS = 0.050, 0.040, 0.025
    for nid in nodos_ruta:
        x, y = pos_sub[nid]
        if nid == hub:
            cn, radio, ec, lw = "#1D9E75", RADIO_BASE, "white", 3.5
        elif nid in nodos_destino:
            cn, radio, ec, lw = COLORES_TIPO.get(G.nodes[nid]["tipo"], "#378ADD"), RADIO_DEST, color_dia, 3.0
        else:
            cn, radio, ec, lw = "#AAAAAA", RADIO_TRANS, "#888888", 1.5

        ax.add_patch(plt.Circle((x, y), radio, color=cn, ec=ec, linewidth=lw, zorder=4))
        cod = G.nodes[nid]["cod"] + ("*" if nid in nodos_transito else "")
        fs_cod = 11 if nid == hub else (10 if nid in nodos_destino else 8)
        ax.text(x, y + 0.008, cod, ha="center", va="center", fontsize=fs_cod, fontweight="bold", color="white", zorder=5)

        nombre = G.nodes[nid]["nombre"]
        if x > 0.78:
            lx, ly, ha = x - 0.015, y + radio + 0.035, "right"
        elif x < 0.22:
            lx, ly, ha = x + 0.015, y + radio + 0.035, "left"
        elif y > 0.78:
            lx, ly, ha = x, y - radio - 0.035, "center"
        else:
            lx, ly, ha = x, y + radio + 0.035, "center"
        peso_f = "bold" if (nid in nodos_destino or nid == hub) else "normal"
        borde_l = color_dia if (nid in nodos_destino or nid == hub) else "#BBBBBB"
        ax.text(lx, ly, nombre, ha=ha, va="center", fontsize=9, fontweight=peso_f, color="#1A1A1A", zorder=6,
                bbox=dict(boxstyle="round,pad=0.35", facecolor="white", edgecolor=borde_l, alpha=0.97, linewidth=1.2))

    for i, nodo in enumerate(secuencia):
        if nodo != hub:
            x, y = pos_sub[nodo]
            ax.text(x + RADIO_DEST + 0.015, y + RADIO_DEST + 0.015, str(i), fontsize=11, fontweight="bold",
                    color=color_dia, zorder=7,
                    bbox=dict(boxstyle="circle,pad=0.18", facecolor="white", edgecolor=color_dia, linewidth=1.5, alpha=0.98))

    ruta_str = " → ".join(G.nodes[n]["cod"] for n in secuencia)
    if len(ruta_str) > 70:
        cods = [G.nodes[n]["cod"] for n in secuencia]
        partes = []
        for i in range(0, len(cods), 7):
            partes.append(" → ".join(cods[i:i+7]))
        ruta_str = "\n".join(partes)

    ax.text(0.01, 0.01, f"Ruta óptima ({criterio}):\n{ruta_str}\nTotal: {costo_total:.1f} {UNIDAD[criterio]}",
            transform=ax.transAxes, fontsize=9, color="#1A1A1A", va="bottom",
            bbox=dict(boxstyle="round,pad=0.6", facecolor="white", edgecolor=color_dia, alpha=0.95, linewidth=1.5))

    leyenda = [
        mpatches.Patch(color="#1D9E75", label=f"Hub: {ATRACTIVOS[hub]['nombre']}"),
        mpatches.Patch(color=color_dia, label="Ruta óptima Dijkstra"),
        mpatches.Patch(color="#AAAAAA", label="Nodo de tránsito (*)"),
        mpatches.Patch(color="#CCCCCC", label="Conexiones no usadas"),
    ]
    ax.legend(handles=leyenda, loc="lower right", fontsize=9, framealpha=0.95, edgecolor="#AAAAAA", fancybox=True)

    ax.set_title(f"DÍA {dia} — {cfg['zona']}\n{NOMBRE_CRIT[criterio]} · {len(destinos)} destinos + hub",
                 fontsize=13, fontweight="bold", color=color_dia, pad=15)
    ax.set_xlim(-0.10, 1.12); ax.set_ylim(-0.08, 1.08)
    ax.axis("off")
    plt.tight_layout(pad=2.0)
    return fig


# ═══════════════════════════════════════════════════════════════════════════
# 10. INTERFAZ DE USUARIO - SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════

st.sidebar.title("🗺️ Rutas Coclé")
st.sidebar.caption("28 Atractivos · Algoritmo de Dijkstra")

seccion = st.sidebar.radio(
    "Ir a:",
    ["🏠 Inicio", "📋 Inventario", "🕸️ Grafo completo", "📅 Ruta por día",
     "🔗 Camino mínimo", "📊 Matrices", "⬇️ Exportar Excel"],
)

# ═══════════════════════════════════════════════════════════════════════════
# 11. FUNCIONES AUXILIARES DE IMÁGENES
# ═══════════════════════════════════════════════════════════════════════════

def imagen_de(cod: str):
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        ruta = os.path.join(CARPETA_IMAGENES, cod + ext)
        if os.path.isfile(ruta):
            return ruta
    return None


def mostrar_imagen(cod: str, nombre: str, ancho=None):
    ruta = imagen_de(cod)
    if ruta:
        st.image(ruta, caption=nombre, use_container_width=(ancho is None), width=ancho)
    else:
        st.markdown(
            f"""<div style="border:1px dashed #999;border-radius:8px;padding:18px;
            text-align:center;color:#888;background:#fafafa;">
            📷 Sin foto<br><small><code>imagenes/{cod}.jpg</code></small>
            </div>""",
            unsafe_allow_html=True,
        )


# ═══════════════════════════════════════════════════════════════════════════
# 12. SECCIONES DE LA APP
# ═══════════════════════════════════════════════════════════════════════════

if seccion == "🏠 Inicio":
    st.title("🗺️ Rutas Turísticas Óptimas — Coclé, Panamá")
    st.markdown(f"""
    **Plataforma con {G.number_of_nodes()} atractivos** y **{G.number_of_edges()} conexiones** viales.
    
    - 🏨 5 hubs (ciudades base)
    - 📅 7 itinerarios diarios
    - ⏱️ Rutas de 6-8 horas
    - 🚗 Salida y regreso al hotel
    """)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("🏛️ Atractivos", G.number_of_nodes())
    col2.metric("🛣️ Conexiones", G.number_of_edges())
    col3.metric("📅 Itinerarios", len(DIAS_CONFIG))
    col4.metric("🏨 Hubs", len([d for d in ATRACTIVOS.values() if d["tipo"] == "Hub/Ciudad"]))

elif seccion == "📋 Inventario":
    st.title("📋 Inventario de Atractivos")
    st.write(f"**Total: {len(ATRACTIVOS)} atractivos**")

    filtro_tipo = st.multiselect("Filtrar por tipo:", sorted({d["tipo"] for d in ATRACTIVOS.values()}))

    for nid, data in sorted(ATRACTIVOS.items()):
        if filtro_tipo and data["tipo"] not in filtro_tipo:
            continue
        color = COLORES_TIPO.get(data["tipo"], "#999999")
        with st.container(border=True):
            col1, col2 = st.columns([1, 3])
            with col1:
                mostrar_imagen(data["cod"], data["nombre"], ancho=150)
            with col2:
                st.markdown(f"**{data['nombre']}** `{data['cod']}` (ID: {nid})")
                st.markdown(f"<span style='background:{color}22;color:{color};padding:2px 10px;border-radius:10px;'>{data['tipo']}</span> 📍 {data['distrito']}", unsafe_allow_html=True)
                st.write(f"⭐ Puntaje: **{data['puntaje']}** | 🔗 Conexiones: **{G.degree(nid)}**")

elif seccion == "🕸️ Grafo completo":
    st.title("Grafo Turístico Completo")
    st.write(f"**28 nodos** · **{G.number_of_edges()} aristas**")
    criterio = st.radio("Criterio:", ["distancia", "tiempo", "costo"], horizontal=True)
    fig = dibujar_grafo_completo(criterio)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

elif seccion == "📅 Ruta por día":
    st.title("📅 Ruta Óptima por Día")

    dia = st.selectbox(
        "Selecciona el día:",
        list(DIAS_CONFIG.keys()),
        format_func=lambda d: f"Día {d} — {DIAS_CONFIG[d]['zona']}",
    )
    criterio = st.radio("Criterio de optimización:", ["distancia", "tiempo", "costo"], horizontal=True)

    cfg = DIAS_CONFIG[dia]
    hub = cfg["hub"]
    destinos = cfg["destinos"]
    orden, costo_total = ruta_optima_dia(destinos, matrices, hub=hub, criterio=criterio)

    if orden is None:
        st.error("No se pudo calcular la ruta para este día.")
    else:
        secuencia = [hub] + orden + [hub]

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🏨 Hotel Base", f"{hub} ({ATRACTIVOS[hub]['cod']})")
        col2.metric("🎯 Atractivos", len(destinos))
        col3.metric("🚗 Tiempo traslado", f"{costo_total:.1f} min")
        col4.metric("📏 Distancia total", f"{sum(matrices['distancia'][secuencia[i]][secuencia[i+1]] for i in range(len(secuencia)-1)):.1f} km")

        st.markdown("### 🚗 Ruta Óptima")
        st.markdown(" → ".join([f"**{ATRACTIVOS[n]['cod']}**" for n in secuencia]))

        # Mapa
        fig = dibujar_grafo_dia(dia, criterio, secuencia, costo_total)
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

        # Tabla de tramos
        st.markdown("### 📋 Detalle de Tramos")
        tramos = []
        for i in range(len(secuencia)-1):
            origen, destino = secuencia[i], secuencia[i+1]
            tramos.append({
                "Tramo": i+1,
                "Origen": f"{origen} ({ATRACTIVOS[origen]['cod']})",
                "Destino": f"{destino} ({ATRACTIVOS[destino]['cod']})",
                "Tiempo (min)": matrices["tiempo"][origen][destino],
                "Distancia (km)": matrices["distancia"][origen][destino],
                "Costo ($)": matrices["costo"][origen][destino],
            })
        st.dataframe(pd.DataFrame(tramos), use_container_width=True, hide_index=True)

elif seccion == "🔗 Camino mínimo":
    st.title("🔗 Camino Mínimo entre Dos Nodos")

    opciones = {f"{d['cod']} — {d['nombre']}": nid for nid, d in sorted(ATRACTIVOS.items())}
    col1, col2 = st.columns(2)
    origen_label = col1.selectbox("Origen:", list(opciones.keys()), index=list(opciones.values()).index(15))
    destino_label = col2.selectbox("Destino:", list(opciones.keys()), index=0)
    criterio = st.radio("Criterio:", ["distancia", "tiempo", "costo"], horizontal=True)

    origen, destino = opciones[origen_label], opciones[destino_label]
    if origen == destino:
        st.warning("Elige dos atractivos distintos.")
    else:
        camino = caminos[criterio][origen][destino]
        if not camino:
            st.error("No existe camino entre esos dos nodos.")
        else:
            valor = matrices[criterio][origen][destino]
            st.success(f"**Ruta:** {' → '.join(G.nodes[n]['cod'] for n in camino)}  —  **Valor total:** {valor:.2f} {UNIDAD[criterio]}")

elif seccion == "📊 Matrices":
    st.title("📊 Matrices de Caminos Mínimos")

    criterio = st.radio("Matriz:", ["distancia", "tiempo", "costo"], horizontal=True)
    nodos = sorted(G.nodes())
    mat = matrices[criterio]

    tabla = []
    for origen in nodos:
        fila = {"": ATRACTIVOS[origen]["cod"]}
        for destino in nodos:
            if origen == destino:
                fila[ATRACTIVOS[destino]["cod"]] = 0
            else:
                val = mat[origen][destino]
                fila[ATRACTIVOS[destino]["cod"]] = "∞" if val is None else round(val, 1)
        tabla.append(fila)

    st.dataframe(tabla, use_container_width=True, height=600, hide_index=True)

elif seccion == "⬇️ Exportar Excel":
    st.title("⬇️ Exportar Matrices a Excel")
    st.write("Descarga el archivo Excel con las matrices de tiempo, distancia y costo.")

    excel_bytes = generar_excel_bytes(matrices, G)

    st.download_button(
        label="⬇️ Descargar Matrices_Dijkstra_Cocle.xlsx",
        data=excel_bytes,
        file_name="Matrices_Dijkstra_Cocle_28Nodos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )

    st.success(f"✅ Archivo generado con {G.number_of_nodes()} nodos y {G.number_of_edges()} aristas")
