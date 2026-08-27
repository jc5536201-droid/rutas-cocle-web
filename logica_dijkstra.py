"""
TESIS: OPTIMIZACIÓN DE RUTAS TURÍSTICAS EN COCLÉ, PANAMÁ
Aplicación web con Streamlit - Versión Corregida
Autor: [Tu nombre]
"""

import streamlit as st
import pandas as pd
import networkx as nx
import heapq
from itertools import permutations
from io import BytesIO
import plotly.graph_objects as go
import plotly.express as px

# ═══════════════════════════════════════════════════════════════════════════
# 1. DATOS DEL SISTEMA TURÍSTICO
# ═══════════════════════════════════════════════════════════════════════════

ATRACTIVOS = {
    1:  {"nombre": "Playa Santa Clara",               "cod": "PSC", "tipo": "Playa",            "puntaje": 27, "distrito": "Antón"},
    2:  {"nombre": "Playa Farallón",                  "cod": "PFA", "tipo": "Playa",            "puntaje": 26, "distrito": "Antón"},
    3:  {"nombre": "Playa El Salado",                 "cod": "PES", "tipo": "Playa",            "puntaje": 21, "distrito": "Aguadulce"},
    4:  {"nombre": "Playa Blanca",                    "cod": "PBL", "tipo": "Playa",            "puntaje": 26, "distrito": "Antón"},
    5:  {"nombre": "Playa Juan Hombrón",              "cod": "PJH", "tipo": "Playa",            "puntaje": 20, "distrito": "Antón"},
    6:  {"nombre": "Mercado Artesanía Valle Antón",   "cod": "MAV", "tipo": "Cultural",         "puntaje": 26, "distrito": "Antón"},
    7:  {"nombre": "Serpentario Maravillas Tropicales","cod": "SMT", "tipo": "Naturaleza",       "puntaje": 24, "distrito": "Antón"},
    8:  {"nombre": "Museo Hermanos Arias Madrid",     "cod": "MHA", "tipo": "Cultural/Hist.",   "puntaje": 25, "distrito": "Penonomé"},
    9:  {"nombre": "P.N. Omar Torrijos",              "cod": "PNT", "tipo": "Parque Nacional",  "puntaje": 22, "distrito": "Penonomé"},
    10: {"nombre": "Sitio Arqueológico El Caño",      "cod": "SAC", "tipo": "Arqueológico",     "puntaje": 26, "distrito": "Natá"},
    11: {"nombre": "Museo Regional Stella Sierra",    "cod": "MSS", "tipo": "Cultural/Hist.",   "puntaje": 22, "distrito": "Aguadulce"},
    12: {"nombre": "Iglesia San Juan Bautista",       "cod": "ISJ", "tipo": "Histórico",        "puntaje": 24, "distrito": "Penonomé"},
    13: {"nombre": "El Chorro Las Yayas",             "cod": "CLY", "tipo": "Cascada",          "puntaje": 25, "distrito": "La Pintada"},
    14: {"nombre": "Balneario Las Mendozas",          "cod": "BLM", "tipo": "Balneario",        "puntaje": 21, "distrito": "Penonomé"},
    15: {"nombre": "Penonomé",                        "cod": "PEN", "tipo": "Hub/Ciudad",       "puntaje": 28, "distrito": "Penonomé"},
    16: {"nombre": "Aguadulce",                       "cod": "AGU", "tipo": "Hub/Ciudad",       "puntaje": 25, "distrito": "Aguadulce"},
    17: {"nombre": "Antón",                           "cod": "ANT", "tipo": "Hub/Ciudad",       "puntaje": 23, "distrito": "Antón"},
    18: {"nombre": "La Pintada",                      "cod": "LAP", "tipo": "Hub/Ciudad",       "puntaje": 22, "distrito": "La Pintada"},
    19: {"nombre": "Natá",                            "cod": "NAT", "tipo": "Hub/Ciudad",       "puntaje": 23, "distrito": "Natá"},
    20: {"nombre": "Parroquia Ntra. Sra. Candelaria", "cod": "PNC", "tipo": "Histórico",        "puntaje": 21, "distrito": "La Pintada"},
    21: {"nombre": "Cerro Gaital",                    "cod": "CGA", "tipo": "Montaña",          "puntaje": 27, "distrito": "Antón"},
}

# DISTANCIAS Y TIEMPOS REALES (medidos en Google Maps)
ARISTAS_RAW = [
    (15, 17,  22.6,  25),  # ⭐ Penonomé → Antón (22.6 km)
    (15, 10,  29.8,  31),
    (15, 16,  50.2,  48),
    (15, 18,  18.4,  24),
    (15,  8,   4.4,   8),
    (15, 12,   4.6,   9),
    (15, 14,   4.6,   9),
    (15, 21,  42.9,  77),
    (17,  1,  23.1,  25),
    (17,  2,  21.4,  25),
    (17,  4,  19.8,  22),
    (17,  5,  18.0,  24),
    (17,  6,  37.6,  59),
    (17,  7,  39.8,  63),
    (17, 21,  40.5,  65),
    ( 1,  2,   6.8,  11),
    ( 2,  4,   8.4,  12),
    ( 4,  5,  24.0,  29),
    ( 1,  5,  27.3,  32),
    ( 6,  7,   2.2,   5),
    ( 6, 21,   2.9,   7),
    (19, 10,  28.7,  37),
    (19, 17,  66.0,  65),
    (19, 16,  19.4,  31),
    (16, 10,  26.8,  27),
    (16, 11,   2.8,   6),
    (16,  3,   2.6,   7),
    (18, 13,  34.2,  54),
    (18, 20,   0.1,   1),
    (18,  9,  41.3,  88),
    ( 9, 13,   7.1,  34),
    ( 9, 10,  44.7,  85),
]

# ARISTAS ADICIONALES PARA COMPLETAR CONEXIONES
ARISTAS_ADICIONALES = [
    (7, 21, 2.5, 5),      # Valle de Antón
    (8, 12, 0.5, 2),      # Penonomé
    (8, 14, 1.0, 3),
    (12, 14, 0.8, 2),
    (13, 20, 0.3, 1),     # La Pintada
    (9, 20, 8.0, 15),
    (11, 3, 1.5, 4),      # Aguadulce
    (17, 18, 35.0, 40),   # Circuito Hubs
    (18, 19, 25.0, 30),
]

# COMBINAR TODAS LAS ARISTAS
ARISTAS = ARISTAS_RAW + ARISTAS_ADICIONALES

# CONFIGURACIÓN DE DÍAS (7 días - rutas circulares)
DIAS_CONFIG = {
    1: {"destinos": [1, 2, 4, 5], "hub": 17, "zona": "Playas de Antón", "color": "#185FA5"},
    2: {"destinos": [6, 7, 21], "hub": 17, "zona": "Valle de Antón", "color": "#854F0B"},
    3: {"destinos": [8, 12, 14], "hub": 15, "zona": "Penonomé Histórico", "color": "#0F6E56"},
    4: {"destinos": [13, 20, 9], "hub": 18, "zona": "Circuito Montañoso", "color": "#534AB7"},
    5: {"destinos": [10, 19], "hub": 19, "zona": "El Caño y Natá", "color": "#993C1D"},
    6: {"destinos": [3, 11], "hub": 16, "zona": "Aguadulce", "color": "#0F6E56"},
    7: {"destinos": [17, 18, 19], "hub": 15, "zona": "Circuito Hubs", "color": "#5B4A00"},
}

# POSICIONES PARA EL MAPA
POSICIONES = {
    15: (0.50, 0.50), 17: (0.74, 0.42), 16: (0.22, 0.20), 18: (0.36, 0.74),
    19: (0.40, 0.26),  1: (0.94, 0.62),  2: (0.94, 0.50),  3: (0.10, 0.10),
     4: (0.92, 0.38),  5: (0.84, 0.26),  6: (0.82, 0.68),  7: (0.94, 0.74),
     8: (0.44, 0.44),  9: (0.18, 0.82), 10: (0.30, 0.30), 11: (0.10, 0.24),
    12: (0.58, 0.44), 13: (0.20, 0.90), 14: (0.50, 0.36), 20: (0.28, 0.84),
    21: (0.80, 0.82),
}

# ═══════════════════════════════════════════════════════════════════════════
# 2. ALGORITMO DIJKSTRA (CORREGIDO)
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_data
def construir_grafo():
    """Construye el grafo con todos los atractivos y aristas."""
    G = nx.Graph()
    
    # Agregar nodos
    for nid, data in ATRACTIVOS.items():
        G.add_node(nid, **data)
    
    # Agregar aristas - VERIFICAR QUE TODAS SE AGREGUEN
    for u, v, dist, tiempo in ARISTAS:
        costo = round(dist * 0.15, 2)
        G.add_edge(u, v, distancia=dist, tiempo=tiempo, costo=costo)
    
    # ✅ FORZAR arista directa 15-17 (por si acaso)
    if not G.has_edge(15, 17):
        G.add_edge(15, 17, distancia=22.6, tiempo=25, costo=3.39)
        print("⚠️ Arista 15-17 forzada")
    
    return G

@st.cache_data
def calcular_matrices(G):
    """Calcula matrices de distancias mínimas usando Dijkstra."""
    nodos = sorted(G.nodes())
    matrices = {}
    
    for criterio in ["distancia", "tiempo", "costo"]:
        mat = {}
        for origen in nodos:
            # Algoritmo de Dijkstra
            INF = float('inf')
            dist = {n: INF for n in G.nodes()}
            dist[origen] = 0
            heap = [(0, origen)]
            visitados = set()
            
            while heap:
                d_u, u = heapq.heappop(heap)
                if u in visitados:
                    continue
                visitados.add(u)
                for v in G.neighbors(u):
                    peso = G[u][v][criterio]
                    alt = dist[u] + peso
                    if alt < dist[v]:
                        dist[v] = alt
                        heapq.heappush(heap, (alt, v))
            
            mat[origen] = {d: round(dist[d], 2) if dist[d] != INF else None for d in nodos}
        matrices[criterio] = mat
    
    return matrices

def ruta_circular_optima(destinos, hub, matrices, criterio="tiempo"):
    """Encuentra la ruta circular óptima usando permutaciones."""
    mat = matrices[criterio]
    mejor_costo = float('inf')
    mejor_orden = None
    mejor_detalle = None
    
    for perm in permutations(destinos):
        secuencia = [hub] + list(perm) + [hub]
        total = 0
        valida = True
        detalle = []
        
        for i in range(len(secuencia) - 1):
            c = mat[secuencia[i]][secuencia[i + 1]]
            if c is None or c == float('inf'):
                valida = False
                break
            total += c
            detalle.append({
                "origen": secuencia[i],
                "destino": secuencia[i+1],
                "costo": c
            })
        
        if valida and total < mejor_costo:
            mejor_costo = total
            mejor_orden = list(perm)
            mejor_detalle = detalle
    
    return mejor_orden, mejor_costo, mejor_detalle

# ═══════════════════════════════════════════════════════════════════════════
# 3. FUNCIONES DE DIAGNÓSTICO
# ═══════════════════════════════════════════════════════════════════════════

def diagnosticar_grafo():
    """Diagnostica problemas en el grafo."""
    G = construir_grafo()
    matrices = calcular_matrices(G)
    
    resultados = {
        "arista_directa": G.has_edge(15, 17),
        "distancia_directa": G[15][17]['distancia'] if G.has_edge(15, 17) else None,
        "tiempo_directo": G[15][17]['tiempo'] if G.has_edge(15, 17) else None,
        "dijkstra_distancia": matrices["distancia"][15][17],
        "dijkstra_tiempo": matrices["tiempo"][15][17],
        "dijkstra_costo": matrices["costo"][15][17],
        "total_aristas": G.number_of_edges(),
        "total_nodos": G.number_of_nodes(),
    }
    
    # Verificar si Dijkstra encontró la ruta óptima
    if resultados["arista_directa"]:
        if resultados["dijkstra_distancia"] == resultados["distancia_directa"]:
            resultados["estado"] = "✅ CORRECTO"
        else:
            resultados["estado"] = f"❌ ERROR: Dijkstra encontró {resultados['dijkstra_distancia']} km, pero la ruta directa es {resultados['distancia_directa']} km"
    
    # Mostrar todas las rutas alternativas
    rutas_alternativas = []
    if G.has_edge(15, 17):
        # Buscar rutas alternativas
        for camino in nx.all_simple_paths(G, 15, 17, cutoff=4):
            if len(camino) > 2:  # Solo rutas con paradas intermedias
                dist = 0
                tiempo = 0
                for i in range(len(camino)-1):
                    dist += G[camino[i]][camino[i+1]]['distancia']
                    tiempo += G[camino[i]][camino[i+1]]['tiempo']
                rutas_alternativas.append({
                    "ruta": camino,
                    "distancia": dist,
                    "tiempo": tiempo
                })
    
    resultados["rutas_alternativas"] = rutas_alternativas
    
    return resultados

# ═══════════════════════════════════════════════════════════════════════════
# 4. INTERFAZ DE USUARIO CON STREAMLIT
# ═══════════════════════════════════════════════════════════════════════════

def main():
    # Configuración de la página
    st.set_page_config(
        page_title="Rutas Turísticas Coclé",
        page_icon="🏝️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Título principal
    st.title("🏝️ Optimización de Rutas Turísticas - Coclé, Panamá")
    st.markdown("---")
    
    # Sidebar con información y diagnóstico
    with st.sidebar:
        st.header("📋 Información")
        st.markdown("""
        **Sistema de optimización de rutas turísticas** utilizando el algoritmo de Dijkstra.
        
        **Características:**
        - 7 días de recorrido
        - Rutas circulares (sale y regresa al hotel)
        - Optimización por tiempo
        - Visualización interactiva
        """)
        
        st.markdown("---")
        
        # 🔍 DIAGNÓSTICO DEL GRAFO
        st.markdown("### 🔍 Diagnóstico del Grafo")
        if st.button("🔍 Verificar Grafo", type="primary"):
            with st.spinner("🔍 Diagnóstico en progreso..."):
                diag = diagnosticar_grafo()
                
                st.markdown("#### 📊 Resultados")
                
                # Métricas principales
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Nodos", diag["total_nodos"])
                with col2:
                    st.metric("Aristas", diag["total_aristas"])
                
                # Verificar arista directa 15-17
                st.markdown("#### 🛣️ Arista 15 (Penonomé) → 17 (Antón)")
                
                if diag["arista_directa"]:
                    st.success(f"✅ Arista directa existe")
                    st.write(f"   📏 Distancia: {diag['distancia_directa']} km")
                    st.write(f"   ⏱ Tiempo: {diag['tiempo_directo']} min")
                else:
                    st.error("❌ ¡Arista directa NO existe!")
                
                # Comparar con Dijkstra
                st.markdown("#### 🔄 Comparación con Dijkstra")
                st.write(f"**Distancia Dijkstra:** {diag['dijkstra_distancia']} km")
                st.write(f"**Tiempo Dijkstra:** {diag['dijkstra_tiempo']} min")
                
                if diag["estado"] == "✅ CORRECTO":
                    st.success("✅ Dijkstra está usando la ruta directa correcta")
                else:
                    st.error(diag["estado"])
                    
                    # Mostrar rutas alternativas
                    if diag["rutas_alternativas"]:
                        st.markdown("#### 🚧 Rutas Alternativas Encontradas")
                        for alt in diag["rutas_alternativas"][:3]:
                            ruta_str = " → ".join(str(n) for n in alt["ruta"])
                            st.write(f"• {ruta_str}")
                            st.write(f"  {alt['distancia']} km, {alt['tiempo']} min")
        
        st.markdown("---")
        st.markdown("**📊 Estadísticas:**")
        st.markdown(f"- **Atractivos:** {len(ATRACTIVOS)}")
        st.markdown(f"- **Conexiones:** {len(ARISTAS)}")
        st.markdown(f"- **Días:** 7")
        
        st.markdown("---")
        st.markdown("**👨‍💻 Tesista:** [Tu nombre]")
    
    # Pestañas para organizar contenido
    tab1, tab2, tab3, tab4 = st.tabs([
        "📅 Plan Semanal",
        "🗺️ Rutas por Día",
        "📊 Matrices Dijkstra",
        "📁 Exportar Datos"
    ])
    
    with tab1:
        st.header("📅 Plan de Viaje Semanal")
        st.markdown("Rutas circulares optimizadas para 7 días")
        
        # Construir grafo y calcular matrices
        with st.spinner("⏳ Calculando rutas óptimas..."):
            G = construir_grafo()
            matrices = calcular_matrices(G)
            
            # Calcular todas las rutas
            resultados = {}
            for dia, config in DIAS_CONFIG.items():
                orden, costo, detalle = ruta_circular_optima(
                    config["destinos"], 
                    config["hub"], 
                    matrices,
                    "tiempo"
                )
                if orden:
                    ruta = [config["hub"]] + orden + [config["hub"]]
                    resultados[dia] = {
                        "ruta": ruta,
                        "tiempo": costo,
                        "destinos": len(config["destinos"]),
                        "detalle": detalle
                    }
        
        # Mostrar resumen en columnas
        cols = st.columns(4)
        for i, (dia, data) in enumerate(resultados.items()):
            with cols[i % 4]:
                config = DIAS_CONFIG[dia]
                st.markdown(f"""
                <div style="
                    background-color: {config['color']}22;
                    padding: 15px;
                    border-radius: 10px;
                    border-left: 5px solid {config['color']};
                    margin-bottom: 10px;
                ">
                    <h4 style="margin: 0;">Día {dia}</h4>
                    <p style="margin: 5px 0; font-size: 0.9em;">{config['zona']}</p>
                    <p style="margin: 5px 0; font-size: 0.8em;">
                        ⏱ {data['tiempo']:.0f} min | 🏠 {ATRACTIVOS[config['hub']]['cod']}
                    </p>
                </div>
                """, unsafe_allow_html=True)
        
        # Tabla detallada
        st.markdown("### 📋 Detalle de Rutas")
        
        df_data = []
        for dia, data in resultados.items():
            config = DIAS_CONFIG[dia]
            ruta_str = " → ".join([f"{n}({ATRACTIVOS[n]['cod']})" for n in data["ruta"]])
            
            # Calcular distancia total
            dist_total = 0
            for i in range(len(data["ruta"])-1):
                dist_total += matrices["distancia"][data["ruta"][i]][data["ruta"][i+1]]
            
            # Verificar si usa la ruta directa (15→17)
            usa_directa = False
            for i in range(len(data["ruta"])-1):
                if data["ruta"][i] == 15 and data["ruta"][i+1] == 17:
                    usa_directa = True
                    break
            
            df_data.append({
                "Día": dia,
                "Zona": config["zona"],
                "Hotel (Hub)": f"{config['hub']} ({ATRACTIVOS[config['hub']]['cod']})",
                "Atractivos": len(data["destinos"]),
                "Ruta": ruta_str,
                "Tiempo (min)": round(data["tiempo"], 1),
                "Distancia (km)": round(dist_total, 1),
                "Usa ruta directa 15→17": "✅" if usa_directa else "❌"
            })
        
        df = pd.DataFrame(df_data)
        st.dataframe(df, use_container_width=True, hide_index=True)
    
    with tab2:
        st.header("🗺️ Visualización de Rutas por Día")
        
        # Selector de día
        dia_seleccionado = st.selectbox(
            "Selecciona un día:",
            options=list(DIAS_CONFIG.keys()),
            format_func=lambda x: f"Día {x} - {DIAS_CONFIG[x]['zona']}"
        )
        
        if dia_seleccionado:
            config = DIAS_CONFIG[dia_seleccionado]
            
            # Obtener ruta optimizada
            G = construir_grafo()
            matrices = calcular_matrices(G)
            orden, costo, detalle = ruta_circular_optima(
                config["destinos"], 
                config["hub"], 
                matrices,
                "tiempo"
            )
            
            if orden:
                ruta = [config["hub"]] + orden + [config["hub"]]
                
                # Mostrar información del día
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("🏠 Hotel Base", f"{ATRACTIVOS[config['hub']]['nombre']}")
                with col2:
                    st.metric("🎯 Atractivos", len(config["destinos"]))
                with col3:
                    st.metric("⏱ Tiempo Total", f"{costo:.1f} min")
                
                # Mostrar ruta
                st.markdown("### 🚗 Ruta Circular Óptima")
                ruta_str = " → ".join([f"**{ATRACTIVOS[n]['nombre']}**" for n in ruta])
                st.markdown(f"{ruta_str}")
                
                # Tabla de tramos
                st.markdown("### 📋 Detalle de Tramos")
                tramos_data = []
                for i in range(len(ruta)-1):
                    origen = ruta[i]
                    destino = ruta[i+1]
                    
                    # Verificar si es la arista directa 15-17
                    es_directa_15_17 = (origen == 15 and destino == 17) or (origen == 17 and destino == 15)
                    
                    tramos_data.append({
                        "Tramo": f"{i+1}",
                        "Origen": f"{origen} ({ATRACTIVOS[origen]['cod']})",
                        "Destino": f"{destino} ({ATRACTIVOS[destino]['cod']})",
                        "Tiempo (min)": matrices["tiempo"][origen][destino],
                        "Distancia (km)": matrices["distancia"][origen][destino],
                        "Directa 15-17": "✅" if es_directa_15_17 else "❌"
                    })
                
                st.dataframe(pd.DataFrame(tramos_data), use_container_width=True, hide_index=True)
                
                # ⚠️ ADVERTENCIA SI NO USA LA RUTA DIRECTA
                usa_directa = False
                for i in range(len(ruta)-1):
                    if ruta[i] == 15 and ruta[i+1] == 17:
                        usa_directa = True
                        break
                
                if not usa_directa and dia_seleccionado in [3, 7]:
                    st.warning("""
                    ⚠️ **¡ADVERTENCIA!** Esta ruta NO está usando la arista directa 15→17 (Penonomé → Antón).
                    
                    Debería usar: **15 → 17 (22.6 km, 25 min)**
                    
                    Pero está usando una ruta alternativa que es más larga.
                    """)
                
                # Mapa interactivo
                st.markdown("### 🗺️ Mapa de la Ruta")
                
                # Crear visualización con Plotly
                fig = go.Figure()
                
                # Dibujar aristas de la ruta
                for i in range(len(ruta)-1):
                    x0, y0 = POSICIONES[ruta[i]]
                    x1, y1 = POSICIONES[ruta[i+1]]
                    
                    # Color especial para la arista 15-17
                    es_arista_especial = (ruta[i] == 15 and ruta[i+1] == 17) or (ruta[i] == 17 and ruta[i+1] == 15)
                    color = "#FF0000" if es_arista_especial else config['color']
                    width = 5 if es_arista_especial else 3
                    
                    fig.add_trace(go.Scatter(
                        x=[x0, x1],
                        y=[y0, y1],
                        mode='lines',
                        line=dict(color=color, width=width),
                        showlegend=False,
                        hoverinfo='none'
                    ))
                
                # Dibujar nodos
                for nid in ruta:
                    x, y = POSICIONES[nid]
                    fig.add_trace(go.Scatter(
                        x=[x],
                        y=[y],
                        mode='markers+text',
                        marker=dict(
                            size=25 if nid in [15, 17] else 20,
                            color='#FF0000' if nid in [15, 17] else config['color'],
                            symbol='star' if nid in [15, 17] else 'circle',
                            line=dict(color='white', width=2)
                        ),
                        text=ATRACTIVOS[nid]['cod'],
                        textposition='top center',
                        name=ATRACTIVOS[nid]['nombre'],
                        hovertemplate=f"<b>{ATRACTIVOS[nid]['nombre']}</b><br>ID: {nid}<extra></extra>"
                    ))
                
                fig.update_layout(
                    height=500,
                    showlegend=False,
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[0, 1]),
                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[0, 1]),
                    plot_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=20, b=20)
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                st.info("""
                📖 **Leyenda del mapa:**
                - ⭐ **Estrella roja:** Nodos 15 (Penonomé) y 17 (Antón)
                - 🔴 **Línea roja:** Arista directa 15-17
                - 🔵 **Círculo azul:** Otros atractivos
                """)
    
    with tab3:
        st.header("📊 Matrices de Dijkstra")
        
        # Selector de criterio
        criterio = st.selectbox(
            "Selecciona el criterio:",
            options=["tiempo", "distancia", "costo"],
            format_func=lambda x: {
                "tiempo": "⏱ Tiempo (minutos)",
                "distancia": "📏 Distancia (km)",
                "costo": "💰 Costo (USD)"
            }[x]
        )
        
        G = construir_grafo()
        matrices = calcular_matrices(G)
        
        # Convertir matriz a DataFrame
        mat = matrices[criterio]
        nodos = sorted(G.nodes())
        
        df_mat = pd.DataFrame(index=nodos, columns=nodos)
        for i in nodos:
            for j in nodos:
                df_mat.loc[i, j] = mat[i][j] if mat[i][j] is not None else "∞"
        
        # Mostrar matriz con estilo
        st.dataframe(
            df_mat,
            use_container_width=True,
            height=600,
            column_config={
                col: st.column_config.TextColumn(f"{col}") for col in df_mat.columns
            }
        )
        
        # Mostrar valores clave
        st.markdown("### 🔍 Valores Clave")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                "15 → 17 (Penonomé → Antón)",
                f"{matrices['distancia'][15][17]} km",
                delta=f"{matrices['tiempo'][15][17]} min"
            )
        
        with col2:
            st.metric(
                "17 → 15 (Antón → Penonomé)",
                f"{matrices['distancia'][17][15]} km",
                delta=f"{matrices['tiempo'][17][15]} min"
            )
        
        with col3:
            st.metric(
                "¿Usa ruta directa?",
                "✅" if matrices['distancia'][15][17] == 22.6 else "❌"
            )
        
        st.info("""
        **📖 Leyenda:**
        - Los valores representan la distancia mínima entre cada par de nodos
        - ∞ significa que no hay conexión posible
        - La diagonal principal siempre es 0 (distancia a sí mismo)
        - **15 → 17 DEBE SER 22.6 km** (ruta directa)
        """)
    
    with tab4:
        st.header("📁 Exportar Datos")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📊 Exportar a Excel")
            st.markdown("Descarga todas las matrices y rutas en un archivo Excel")
            
            if st.button("📥 Generar y Descargar Excel", type="primary"):
                with st.spinner("⏳ Generando archivo..."):
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment
                        
                        wb = Workbook()
                        
                        # Hoja de rutas
                        ws = wb.active
                        ws.title = "Rutas"
                        
                        # Título
                        ws['A1'] = "RUTAS TURÍSTICAS - COCLÉ, PANAMÁ"
                        ws['A1'].font = Font(bold=True, size=14)
                        ws.merge_cells('A1:F1')
                        
                        # Encabezados
                        headers = ["Día", "Zona", "Hub", "Ruta", "Tiempo (min)", "Distancia (km)"]
                        for i, h in enumerate(headers, 1):
                            cell = ws.cell(row=2, column=i, value=h)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                        
                        # Datos
                        G = construir_grafo()
                        matrices = calcular_matrices(G)
                        
                        row = 3
                        for dia, config in DIAS_CONFIG.items():
                            orden, costo, _ = ruta_circular_optima(
                                config["destinos"], 
                                config["hub"], 
                                matrices,
                                "tiempo"
                            )
                            if orden:
                                ruta = [config["hub"]] + orden + [config["hub"]]
                                ruta_str = " → ".join(str(n) for n in ruta)
                                
                                # Calcular distancia
                                dist_total = 0
                                for i in range(len(ruta)-1):
                                    dist_total += matrices["distancia"][ruta[i]][ruta[i+1]]
                                
                                ws.cell(row=row, column=1, value=dia)
                                ws.cell(row=row, column=2, value=config["zona"])
                                ws.cell(row=row, column=3, value=f"{config['hub']} ({ATRACTIVOS[config['hub']]['cod']})")
                                ws.cell(row=row, column=4, value=ruta_str)
                                ws.cell(row=row, column=5, value=round(costo, 1))
                                ws.cell(row=row, column=6, value=round(dist_total, 1))
                                row += 1
                        
                        # Ajustar anchos
                        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                            ws.column_dimensions[col].width = 20
                        
                        # Guardar en buffer
                        buffer = BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        
                        # Descargar
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=buffer,
                            file_name="rutas_turisticas.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        st.success("✅ Archivo Excel generado exitosamente!")
                        
                    except Exception as e:
                        st.error(f"❌ Error al generar Excel: {e}")
        
        with col2:
            st.markdown("### 📋 Exportar a CSV")
            st.markdown("Descarga las rutas en formato CSV")
            
            # Preparar datos para CSV
            G = construir_grafo()
            matrices = calcular_matrices(G)
            
            csv_data = []
            for dia, config in DIAS_CONFIG.items():
                orden, costo, _ = ruta_circular_optima(
                    config["destinos"], 
                    config["hub"], 
                    matrices,
                    "tiempo"
                )
                if orden:
                    ruta = [config["hub"]] + orden + [config["hub"]]
                    ruta_str = " → ".join(str(n) for n in ruta)
                    
                    # Calcular distancia
                    dist_total = 0
                    for i in range(len(ruta)-1):
                        dist_total += matrices["distancia"][ruta[i]][ruta[i+1]]
                    
                    csv_data.append({
                        "Día": dia,
                        "Zona": config["zona"],
                        "Hub": config["hub"],
                        "Ruta": ruta_str,
                        "Tiempo (min)": round(costo, 1),
                        "Distancia (km)": round(dist_total, 1)
                    })
            
            df_csv = pd.DataFrame(csv_data)
            
            # Botón de descarga CSV
            csv = df_csv.to_csv(index=False)
            st.download_button(
                label="📥 Descargar CSV",
                data=csv,
                file_name="rutas_turisticas.csv",
                mime="text/csv",
                use_container_width=True
            )

# ═══════════════════════════════════════════════════════════════════════════
# 5. EJECUCIÓN
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
